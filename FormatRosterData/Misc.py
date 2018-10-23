import os
import openpyxl
import TM_CommonPy as TM

def GetPos(vCell,iColAdjustment=0):
    return openpyxl.utils.get_column_letter(openpyxl.utils.column_index_from_string(vCell.column)+iColAdjustment)+str(vCell.row)

def SplitName(vOldSheet,vNewSheet):
    iMaxCol = len(vNewSheet['1'])
    bSuccess = True
    #---Determine Name Column and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "name" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        return False
    #---
    for vCell in vOldSheet['B']:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        cSplitString = vCell.value.strip().split(None,1) #split None splits at first whitespace, a necessary bugfix
        try:
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = cSplitString[0]
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+2)+str(vCell.row)] = cSplitString[1]
        except IndexError:
            bSuccess = False
    return bSuccess

def SplitTown(vOldSheet,vNewSheet):
    iMaxCol = len(vNewSheet['1'])
    bSuccess = True
    #---Determine Name Column and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "hometown" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("Could not find HOMETOWN Column and Row")
        return False
    #---
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        cSplitString = vCell.value.split(", ")
        try:
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = cSplitString[0]
            vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+2)+str(vCell.row)] = cSplitString[1].split("/")[0].strip()
        except:
            bSuccess=False
            raise
    return bSuccess

def ConvertDateToHeight(vDate):
    #---Filter
    if vDate is None:
        return ""
    #---
    cTemp = str(vDate).split("-")
    return cTemp[1]+"\'"+cTemp[2].split(None)[0]+"\""

def TranslateHeight(vOldSheet,vNewSheet):
    iMaxCol = len(vNewSheet['1'])
    bSuccess = True
    #---Determine Height Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "ht." in vCell.value.lower() or "height" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("Could not find Height Column and Row")
        return False
    #---
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = ConvertDateToHeight(vCell.value)
    return bSuccess

def GetWeight(vOldSheet,vNewSheet):
    iMaxCol = len(vNewSheet['1'])
    bSuccess = True
    #---Determine Weight Col and Row
    for vCell in (vOldSheet['1']+vOldSheet['2']):
        try:
            if "wt." in vCell.value.lower() or "weight" in vCell.value.lower():
                iRow = vCell.row
                sColumn = vCell.column
                break
        except (TypeError, AttributeError):
            pass
    else:
        print("Could not find Weight Column and Row")
        return False
    #---
    for vCell in vOldSheet[sColumn]:
        #-Skip past header
        if vCell.row <= iRow:
            continue
        #-
        vNewSheet[openpyxl.utils.get_column_letter(iMaxCol+1)+str(vCell.row)] = vCell.value
    return bSuccess

def AppendOldSheet(vOldSheet,vNewSheet):
    iMaxCol = len(vNewSheet['1'])
    bSuccess = True
    for cColumn in vOldSheet.iter_cols():
        for vCell in cColumn:
            vNewSheet[GetPos(vCell,iColAdjustment=iMaxCol+1)] = vCell.value
    return bSuccess
