import os
import openpyxl
import TM_CommonPy as TM
import requests

def DigForText(vElem):
    for i in range(15):
        if not vElem.text is None:
            break;
        vElem = vElem[0]
    else:
        return ""
    return vElem.text

def GetRoster(sURL):
    #---Get RosterTable
    vRosterPage = requests.get(sURL)
    tree = lxml.html.fromstring(vRosterPage.content)
    vRosterTableHeader = tree.xpath('//thead[@class="Table2__sub-header Table2__thead"]')[0]
    vRosterTable = tree.xpath('//tbody[@class="Table2__tbody"]')[0]
    vRosterTitle = tree.xpath('//h1[@class="headline__h1 dib"]')
    sRosterTitle = vRosterTitle[0].text.replace(" ","")
    #---Convert vRosterTable to openpyxl doc
    vWorkbook = openpyxl.Workbook()
    vSheet = vWorkbook.active
    for iCol, vItem in enumerate(vRosterTableHeader[0]):
        vSheet[openpyxl.utils.get_column_letter(iCol+1)+str(1)] = DigForText(vItem)
    for iRow, vRow in enumerate(vRosterTable):
        for iCol, vItem in enumerate(vRow):
            vSheet[openpyxl.utils.get_column_letter(iCol+1)+str(iRow+1+1)] = DigForText(vItem) #xlsx iCol and iRow start index at 1. Row gets another +1 for header.
    return vWorkbook
    vWorkbook.save("ScrapedData_"+sRosterTitle+".xlsx")
